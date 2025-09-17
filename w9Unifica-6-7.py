import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Arquivos
# -----------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_NOMES_COMPLETA2.xlsx"

# -----------------------------
# Carregar delegados e férias
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()

# Cria dicionário código → nome
mapa_codigos = dict(zip(df_delegados["Código"], df_delegados["Nome"]))

# Converte colunas de férias
for col in ["Inicio Férias 1", "Término Férias 1", "Inicio Férias 2", "Término Férias 2"]:
    df_delegados[col] = pd.to_datetime(df_delegados[col], errors="coerce")

# Monta dicionário de férias
ferias = {}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    if pd.notna(row["Inicio Férias 1"]) and pd.notna(row["Término Férias 1"]):
        periodos.append((row["Inicio Férias 1"].date(), row["Término Férias 1"].date()))
    if pd.notna(row["Inicio Férias 2"]) and pd.notna(row["Término Férias 2"]):
        periodos.append((row["Inicio Férias 2"].date(), row["Término Férias 2"].date()))
    if periodos:
        ferias[codigo] = periodos

# Lista de delegados para rodízio nos finais de semana (códigos 3–22)
delegados_rodizio = df_delegados[df_delegados["Código"].astype(int).between(3, 22)]
lista_delegados = list(zip(delegados_rodizio["Código"], delegados_rodizio["Nome"]))

# -----------------------------
# Carregar planilha da escala
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active

col_data = 1    # coluna A
col_diurno = 3  # coluna C
col_noturno = 4 # coluna D

# -----------------------------
# Padrão de preenchimento para plantões regulares
# -----------------------------
padrao = [("1", "2"), ("", "1"), ("2", ""), ("", ""), ("", "")]
indice = 0

# -----------------------------
# Primeira etapa: preencher plantões regulares
# -----------------------------
for row in range(2, ws.max_row + 1):
    cell_data = ws.cell(row=row, column=col_data).value
    if not cell_data:
        continue

    data = parse_data(cell_data)
    if not data:
        continue

    # pega os próximos códigos do padrão
    c_val, d_val = padrao[indice]
    indice = (indice + 1) % len(padrao)

    # substitui código por nome
    c_nome = mapa_codigos.get(c_val, "") if c_val else ""
    d_nome = mapa_codigos.get(d_val, "") if d_val else ""

    # escreve provisoriamente
    ws.cell(row=row, column=col_diurno, value=c_nome)
    ws.cell(row=row, column=col_noturno, value=d_nome)

    # verifica férias do diurno
    if c_val and c_val in ferias:
        for inicio, fim in ferias[c_val]:
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_diurno, value="")

    # verifica férias do noturno
    if d_val and d_val in ferias:
        for inicio, fim in ferias[d_val]:
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_noturno, value="")

# -----------------------------
# Segunda etapa: preencher finais de semana noturnos em branco
# -----------------------------
idx = 0
for row in range(2, ws.max_row + 1):
    cell_data = ws.cell(row=row, column=col_data).value
    data = parse_data(cell_data)
    if not data:
        continue

    # sexta (4), sábado (5) ou domingo (6)
    if data.weekday() in (4, 5, 6):
        cell_noturno = ws.cell(row=row, column=col_noturno)
        if cell_noturno.value in (None, "", " "):
            tentativas = 0
            while tentativas < len(lista_delegados):
                codigo, nome = lista_delegados[idx]
                # verifica se está em férias
                em_ferias = False
                if codigo in ferias:
                    for inicio, fim in ferias[codigo]:
                        if inicio <= data <= fim:
                            em_ferias = True
                            break
                if not em_ferias:
                    cell_noturno.value = nome
                    idx = (idx + 1) % len(lista_delegados)
                    break
                else:
                    idx = (idx + 1) % len(lista_delegados)
                    tentativas += 1

# -----------------------------
# Salvar planilha final
# -----------------------------
wb.save(saida_final)
print(f"✅ Escala final completa salva em: {saida_final}")
