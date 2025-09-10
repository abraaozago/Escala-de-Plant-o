from openpyxl import load_workbook

# Abre a planilha mantendo formatação
arquivo = "ESCALA 2º Semestre 2025.xlsx"
wb = load_workbook(arquivo)
ws = wb.active  # usa a primeira aba, troque por wb["NomeDaAba"] se necessário

# Lista de dias que devem ser tratados
dias_fds = ["sexta", "sábado", "sabado", "domingo"]

# Contador de códigos cíclico
codigo_atual = 3

# Percorre as linhas (começando da 2ª, assumindo cabeçalho na 1ª)
for row in range(2, ws.max_row + 1):
    dia_semana = ws.cell(row=row, column=2).value  # Coluna B
    celula_d = ws.cell(row=row, column=4)  # Coluna D

    if dia_semana:
        dia_semana_lower = str(dia_semana).strip().lower()
        if any(dia in dia_semana_lower for dia in dias_fds):
            if not celula_d.value:  # Se estiver vazio
                celula_d.value = codigo_atual
                codigo_atual += 1
                if codigo_atual > 23:  # volta para 3 quando passa do 23
                    codigo_atual = 3

# Salva em novo arquivo
novo_arquivo = "ESCALA_2º_Semestre_2025_com_codigos_ciclicos.xlsx"
wb.save(novo_arquivo)

print(f"✅ Planilha gerada com códigos cíclicos de 3 a 23 preenchidos: {novo_arquivo}")
