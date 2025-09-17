import pandas as pd
from openpyxl import load_workbook

# -----------------------------
# Caminhos dos arquivos
# -----------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL.xlsx"
saida_nomes = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_NOMES.xlsx"

# -----------------------------
# Lê delegados e cria mapa de códigos para nomes
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()

# supondo que a planilha tenha coluna "Nome"
mapa_codigos = dict(zip(df_delegados["Código"], df_delegados["Nome"]))

# -----------------------------
# Abre a escala final
# -----------------------------
wb = load_workbook(arquivo_escala_final)
ws = wb.active  # assume aba principal

col_diurno = 3   # coluna C
col_noturno = 4  # coluna D

# -----------------------------
# Substituir códigos pelos nomes
# -----------------------------
for row in range(2, ws.max_row + 1):  # pula cabeçalho
    for col in [col_diurno, col_noturno]:
        valor = ws.cell(row=row, column=col).value
        if valor is None:
            continue
        valor_str = str(valor).strip()
        if valor_str in mapa_codigos:
            ws.cell(row=row, column=col, value=mapa_codigos[valor_str])

# -----------------------------
# Salvar nova planilha
# -----------------------------
wb.save(saida_nomes)
print(f"✅ Escala com nomes salva em: {saida_nomes}")
