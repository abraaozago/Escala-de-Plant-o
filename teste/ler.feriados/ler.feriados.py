import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, date

# ======================================================
# CONFIGURAÇÕES
# ======================================================

ARQ_FERIADOS = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados2.xlsx"
ABA_FERIADOS = "Feriados"
ABA_DELEGADOS = "Delegados"

ARQ_ESCALA = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
ARQ_SAIDA = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025 - FERIADOS.xlsx"

# Colunas da escala
COL_DATA = "A"
COL_DIURNO = "C"
COL_NOTURNO = "D"

# Cor verde (para feriados)
fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# ======================================================
# Funções auxiliares
# ======================================================

def to_date_or_none(v):
    """Converte valores vindos do excel para date ou retorna None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # tentar parse de string (caso tenha sido salvo como texto)
    try:
        # formatos comuns: YYYY-MM-DD ou DD/MM/YYYY
        s = str(v).strip()
        if "-" in s:
            # tenta ISO
            return datetime.fromisoformat(s).date()
        if "/" in s:
            parts = s.split("/")
            if len(parts) == 3:
                # assumir DD/MM/YYYY
                d, m, y = parts
                return date(int(y), int(m), int(d))
    except Exception:
        pass
    return None

def date_in_any_range(d, ranges):
    """Retorna True se date d (date) estiver dentro de qualquer tupla (start,end) em ranges.
       start/end inclusive. Se start ou end for None ignora essa tupla.
    """
    if d is None:
        return False
    for (s, e) in ranges:
        if s is None or e is None:
            continue
        if s <= d <= e:
            return True
    return False

# ======================================================
# 1) Carregar feriados (convertendo para date)
# ======================================================

wb_f = openpyxl.load_workbook(ARQ_FERIADOS, data_only=True)
ws_f = wb_f[ABA_FERIADOS]

feriados = set()

for row in ws_f.iter_rows(min_row=2, max_col=1):
    cell = row[0].value
    if isinstance(cell, datetime):
        feriados.add(cell.date())
    elif isinstance(cell, date):
        feriados.add(cell)

print("Feriados carregados:", feriados)

# ======================================================
# 2) Carregar delegados plantonistas e períodos de férias
#    - Colunas esperadas na aba "Delegados":
#      A = Nome
#      B = Título (deve ser "plantão")
#      F = início férias 1
#      G = fim    férias 1
#      H = início férias 2
#      I = fim    férias 2
# ======================================================

ws_delegados = wb_f[ABA_DELEGADOS]

delegados_plantao = []          # lista na ordem encontrada
delegados_ferias_ranges = {}    # nome -> list de (start_date, end_date)

# iterar até a coluna I (9 colunas) para garantir acesso a F-I
for row in ws_delegados.iter_rows(min_row=2, max_col=9):
    nome = row[0].value           # A
    titulo = row[1].value         # B

    if nome is None:
        continue

    # normalizar nome
    nome_str = str(nome).strip()

    if titulo and str(titulo).strip().lower() == "plantão":
        delegados_plantao.append(nome_str)

        # ler períodos de férias: F,G e H,I = indices 5,6 e 7,8
        f_inicio_1 = to_date_or_none(row[5].value)  # F
        f_fim_1    = to_date_or_none(row[6].value)  # G
        f_inicio_2 = to_date_or_none(row[7].value)  # H
        f_fim_2    = to_date_or_none(row[8].value)  # I

        ranges = []
        if f_inicio_1 and f_fim_1:
            # garantir ordem correta
            if f_inicio_1 <= f_fim_1:
                ranges.append((f_inicio_1, f_fim_1))
            else:
                ranges.append((f_fim_1, f_inicio_1))
        if f_inicio_2 and f_fim_2:
            if f_inicio_2 <= f_fim_2:
                ranges.append((f_inicio_2, f_fim_2))
            else:
                ranges.append((f_fim_2, f_inicio_2))

        delegados_ferias_ranges[nome_str] = ranges

# Para qualquer delegado plantão sem entrada explícita em dict, garantir chave vazia
for d in delegados_plantao:
    delegados_ferias_ranges.setdefault(d, [])

print("Delegados plantonistas (ordem lida):", delegados_plantao)
print("Períodos de férias por delegado:")
for nome, ranges in delegados_ferias_ranges.items():
    print(" -", nome, ":", ranges)

# ======================================================
# 3) Abrir escala e usar a primeira aba
# ======================================================

wb_e = openpyxl.load_workbook(ARQ_ESCALA)
ws_e = wb_e[wb_e.sheetnames[0]]

print("Usando aba da escala:", wb_e.sheetnames[0])

# ======================================================
# 4) Aplicar cores dos feriados e vésperas (colunas C e D)
# ======================================================

for row in ws_e.iter_rows(min_row=2):
    cell_data = row[0]  # coluna A
    data = cell_data.value

    # converter corretamente para date
    if isinstance(data, datetime):
        dia = data.date()
    elif isinstance(data, date):
        dia = data
    else:
        continue

    # DIA DO FERIADO → DIURNO + NOTURNO verde
    if dia in feriados:
        ws_e[f"{COL_DIURNO}{cell_data.row}"].fill = fill_green
        ws_e[f"{COL_NOTURNO}{cell_data.row}"].fill = fill_green

    # VÉSPERA DO FERIADO → apenas NOTURNO
    if (dia + timedelta(days=1)) in feriados:
        ws_e[f"{COL_NOTURNO}{cell_data.row}"].fill = fill_green

# ======================================================
# 5) Preencher PLANTÕES NOTURNOS (ciclo fixo de 4 dias)
#    - ciclo sempre tem 4 posições
#    - se menos de 4 delegados, posições extras são vazias
#    - se delegado está de férias naquele dia, sua posição vira vazia
#    - o ciclo avança independente de férias
# ======================================================

# Construir ciclo fixo de 4 dias a partir da lista lida (ordem importa)
ciclo = []
for i in range(4):
    if i < len(delegados_plantao):
        ciclo.append(delegados_plantao[i])
    else:
        ciclo.append("")   # dias excedentes são vazios

print("Ciclo final utilizado (4 dias):", ciclo)

indice_ciclo = 0

for row in ws_e.iter_rows(min_row=2):
    cell_data = row[0]
    data = cell_data.value

    if isinstance(data, datetime):
        dia = data.date()
    elif isinstance(data, date):
        dia = data
    else:
        continue

    nome_no_ciclo = ciclo[indice_ciclo]

    # decidir valor a escrever na célula do noturno
    valor_para_escrever = ""

    if nome_no_ciclo:
        # checar se esse delegado está de férias neste dia
        ranges = delegados_ferias_ranges.get(nome_no_ciclo, [])
        if not date_in_any_range(dia, ranges):
            valor_para_escrever = nome_no_ciclo
        else:
            # delegado em férias: deixa vazio
            valor_para_escrever = ""

    else:
        # posição do ciclo já é vazia
        valor_para_escrever = ""

    ws_e[f"{COL_NOTURNO}{cell_data.row}"] = valor_para_escrever

    # avançar ciclo (sempre)
    indice_ciclo = (indice_ciclo + 1) % 4

# ======================================================
# 6) Salvar arquivo novo
# ======================================================

wb_e.save(ARQ_SAIDA)

print("\nArquivo salvo com sucesso:")
print(ARQ_SAIDA)
