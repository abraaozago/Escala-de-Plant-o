import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, date
from collections import defaultdict

# ======================================================
# CONFIGURAÇÕES
# ======================================================
ARQ_FERIADOS = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados2.xlsx"
ABA_FERIADOS = "Feriados"
ABA_DELEGADOS = "Delegados"

ARQ_ESCALA = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
ARQ_SAIDA = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025 - COMPLETA.xlsx"

COL_DATA = "A"
COL_DIURNO = "C"
COL_NOTURNO = "D"

fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# ======================================================
# FUNÇÕES AUXILIARES
# ======================================================

def carregar_feriados(wb, aba):
    ws = wb[aba]
    feriados_set = set()
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0].value
        if isinstance(cell, datetime):
            feriados_set.add(cell.date())
        elif isinstance(cell, date):
            feriados_set.add(cell)
    return feriados_set

def carregar_delegados(wb, aba):
    ws = wb[aba]
    plantonistas, expedientes = [], []
    ferias_plantao, ferias_exp = {}, {}

    for row in ws.iter_rows(min_row=2):
        nome, tipo = row[0].value, row[1].value
        if not nome or not tipo:
            continue
        tipo_lower = str(tipo).strip().lower()
        f1_ini, f1_fim = row[5].value, row[6].value
        f2_ini, f2_fim = row[7].value, row[8].value

        periodos = []
        for ini, fim in [(f1_ini, f1_fim), (f2_ini, f2_fim)]:
            if isinstance(ini, (date, datetime)) and isinstance(fim, (date, datetime)):
                ini_dt = ini.date() if isinstance(ini, datetime) else ini
                fim_dt = fim.date() if isinstance(fim, datetime) else fim
                periodos.append((ini_dt, fim_dt))

        if tipo_lower in ["plantão", "plantao"]:
            plantonistas.append(nome)
            ferias_plantao[nome] = periodos
        elif tipo_lower == "expediente":
            expedientes.append(nome)
            ferias_exp[nome] = periodos
    return plantonistas, ferias_plantao, expedientes, ferias_exp

def esta_de_ferias(nome, data, ferias_dict):
    if nome not in ferias_dict:
        return False
    return any(ini <= data <= fim for ini, fim in ferias_dict[nome])

def marcar_feriados(ws, feriados):
    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        if data_simplificada in feriados:
            ws[f"{COL_DIURNO}{data_cell.row}"].fill = fill_green
            ws[f"{COL_NOTURNO}{data_cell.row}"].fill = fill_green
        if (data_simplificada + timedelta(days=1)) in feriados:
            ws[f"{COL_NOTURNO}{data_cell.row}"].fill = fill_green

def registrar_noturnos(ws):
    noturno_realizado = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        nome_noturno = ws[f"{COL_NOTURNO}{data_cell.row}"].value
        if nome_noturno:
            noturno_realizado[nome_noturno].append(data_simplificada)
    return noturno_realizado

def preencher_noturnos_plantonistas(ws, ciclo, ferias_dict):
    idx = 0
    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        col_noturno = ws[f"{COL_NOTURNO}{data_cell.row}"]
        nome = ciclo[idx]
        if nome and not esta_de_ferias(nome, data_simplificada, ferias_dict):
            col_noturno.value = nome
        else:
            col_noturno.value = None
        idx = (idx + 1) % len(ciclo)

def preencher_noturnos_expedientes(ws, expedientes, ferias_dict):
    if not expedientes:
        return
    idx_exp = 0
    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        col_noturno = ws[f"{COL_NOTURNO}{data_cell.row}"]
        if not col_noturno.value:
            tentativas = 0
            while tentativas < len(expedientes):
                nome_exp = expedientes[idx_exp]
                if not esta_de_ferias(nome_exp, data_simplificada, ferias_dict):
                    col_noturno.value = nome_exp
                    break
                idx_exp = (idx_exp + 1) % len(expedientes)
                tentativas += 1
            idx_exp = (idx_exp + 1) % len(expedientes)

def preencher_diurnos(ws, expedientes, ferias_dict, noturno_realizado):
    if not expedientes:
        return
    idx = 0
    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        col_diurno = ws[f"{COL_DIURNO}{data_cell.row}"]
        tentativas = 0
        while tentativas < len(expedientes):
            nome = expedientes[idx]
            if esta_de_ferias(nome, data_simplificada, ferias_dict):
                idx = (idx + 1) % len(expedientes)
                tentativas += 1
                continue
            dias_folga = [data_simplificada - timedelta(days=1),
                          data_simplificada - timedelta(days=2)]
            if any(d in noturno_realizado.get(nome, []) for d in dias_folga):
                idx = (idx + 1) % len(expedientes)
                tentativas += 1
                continue
            col_diurno.value = nome
            idx = (idx + 1) % len(expedientes)
            break

def gerar_resumo(ws):
    if "Resumo Plantões" in ws.parent.sheetnames:
        ws.parent.remove(ws.parent["Resumo Plantões"])
    ws_resumo = ws.parent.create_sheet("Resumo Plantões")
    ws_resumo.append(["Delta", "Diurno Semana", "Diurno FDS", "Noturno Semana", "Noturno FDS"])
    quantidade = defaultdict(lambda: {"diurno_semana":0,"diurno_fds":0,"noturno_semana":0,"noturno_fds":0})

    for row in ws.iter_rows(min_row=2):
        data_cell = row[0]
        data = data_cell.value
        if not isinstance(data, (datetime, date)):
            continue
        data_simplificada = data.date() if isinstance(data, datetime) else data
        dia_semana = data_simplificada.weekday()
        is_fds = dia_semana >= 5 or data_simplificada in feriados

        # Diurno
        diurno_nome = ws[f"{COL_DIURNO}{data_cell.row}"].value
        if diurno_nome:
            if is_fds:
                quantidade[diurno_nome]["diurno_fds"] += 1
            else:
                quantidade[diurno_nome]["diurno_semana"] += 1

        # Noturno
        noturno_nome = ws[f"{COL_NOTURNO}{data_cell.row}"].value
        if noturno_nome:
            is_fds_noturno = dia_semana >= 4 or data_simplificada in feriados or (data_simplificada + timedelta(days=1)) in feriados
            if is_fds_noturno:
                quantidade[noturno_nome]["noturno_fds"] += 1
            else:
                quantidade[noturno_nome]["noturno_semana"] += 1

    for delegado, cont in quantidade.items():
        ws_resumo.append([
            delegado,
            cont['diurno_semana'],
            cont['diurno_fds'],
            cont['noturno_semana'],
            cont['noturno_fds']
        ])

# ======================================================
# EXECUÇÃO PRINCIPAL
# ======================================================

# Carregar feriados e delegados
wb_feriados = openpyxl.load_workbook(ARQ_FERIADOS, data_only=True)
feriados = carregar_feriados(wb_feriados, ABA_FERIADOS)
plantonistas, ferias_plantao, expedientes, ferias_exp = carregar_delegados(wb_feriados, ABA_DELEGADOS)

# Abrir escala
wb_escala = openpyxl.load_workbook(ARQ_ESCALA)
ws_escala = wb_escala[wb_escala.sheetnames[0]]

# Marcar feriados
marcar_feriados(ws_escala, feriados)

# Preencher NOTURNOS
ciclo_noturno = plantonistas[:4]
while len(ciclo_noturno) < 4:
    ciclo_noturno.append(None)
preencher_noturnos_plantonistas(ws_escala, ciclo_noturno, ferias_plantao)
preencher_noturnos_expedientes(ws_escala, expedientes, ferias_exp)

# Preencher DIURNOS (somente expedientes)
noturno_realizado = registrar_noturnos(ws_escala)
preencher_diurnos(ws_escala, expedientes, ferias_exp, noturno_realizado)

# Gerar resumo de plantões
gerar_resumo(ws_escala)

# Salvar arquivo final
wb_escala.save(ARQ_SAIDA)
print("\nArquivo salvo com sucesso! Escala completa e resumo gerado.")
