import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime, timedelta

def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        try:
            return pd.to_datetime(cell).date()
        except:
            return None

def in_ferias(nome, data, ferias_map):
    return any(i <= data <= f for i, f in ferias_map.get(nome, []))

# Arquivos
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados2.xlsx"
arquivo_base      = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final       = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL.xlsx"

# Carregar delegados
df_delegados = pd.read_excel(arquivo_delegados)

# Plantonistas
df_plantonistas = df_delegados[df_delegados.iloc[:,1].astype(str).str.lower().str.contains("plantão")]
plantonistas = df_plantonistas.iloc[:,0].dropna().astype(str).str.strip().tolist()

# Expedientes
df_expedientes = df_delegados[df_delegados.iloc[:,1].astype(str).str.lower().str.contains("expediente")]
expedientes = df_expedientes.iloc[:,0].dropna().astype(str).str.strip().tolist()

# Férias
def map_ferias(df):
    ferias = {}
    for _, row in df.iterrows():
        nome = str(row.iloc[0]).strip()
        periodos = []
        for i in [1, 2]:
            ini_col = f"Inicio Férias {i}"
            fim_col = f"Término Férias {i}"
            if ini_col in df.columns and fim_col in df.columns:
                ini = row.get(ini_col)
                fim = row.get(fim_col)
                if pd.notna(ini) and pd.notna(fim):
                    try:
                        periodos.append((pd.to_datetime(ini).date(), pd.to_datetime(fim).date()))
                    except:
                        pass
        if periodos:
            ferias[nome] = periodos
    return ferias

ferias_plant = map_ferias(df_plantonistas)
ferias_exped = map_ferias(df_expedientes)

# Carregar matriz
wb_base = load_workbook(arquivo_base)
ws_base = wb_base.active
COL_DATA, COL_DIURNO, COL_NOTURNO = 1, 3, 4

# Criar nova planilha copiando layout
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Escala"

for row in ws_base.iter_rows():
    for cell in row:
        new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)

# Copiar dimensões e mesclagens
for col_letter, dim in ws_base.column_dimensions.items():
    ws_out.column_dimensions[col_letter].width = dim.width
for row_num, dim in ws_base.row_dimensions.items():
    ws_out.row_dimensions[row_num].height = dim.height
for merged_range in ws_base.merged_cells.ranges:
    ws_out.merge_cells(str(merged_range))
ws_out.freeze_panes = ws_base.freeze_panes

# Extrair datas (linha, data)
datas = []
for r in range(2, ws_base.max_row+1):
    d = parse_data(ws_base.cell(row=r, column=COL_DATA).value)
    if d:
        datas.append((r, d))

first_date = datas[0][1]
last_date  = datas[-1][1]

# Ciclo 12x24 – 12x72 para plantonistas
ocorrencias_por_nome = {nome: set() for nome in plantonistas}
for i, nome in enumerate(plantonistas):
    offset_days = i % 5
    current_date = first_date + timedelta(days=offset_days)
    turno_diurno = True
    while current_date <= last_date:
        period = "diurno" if turno_diurno else "noturno"
        ocorrencias_por_nome[nome].add((current_date, period))
        step_days = 1 if turno_diurno else 4
        current_date += timedelta(days=step_days)
        turno_diurno = not turno_diurno

# Contagem (para resumo)
contagem = {nome: {"Diurno Semana":0,"Noturno Semana":0,"Diurno FimSemana":0,"Noturno FimSemana":0}
            for nome in plantonistas + expedientes}

# Etapa 1: Plantonistas
start_rr = 0
for idx, (row, data) in enumerate(datas):
    # Diurno
    for j in range(len(plantonistas)):
        nome = plantonistas[(start_rr+j) % len(plantonistas)]
        if (data,"diurno") in ocorrencias_por_nome[nome]:
            if not in_ferias(nome, data, ferias_plant):
                ws_out.cell(row=row, column=COL_DIURNO, value=nome)
                dw = data.weekday()
                if dw < 5: contagem[nome]["Diurno Semana"] += 1
                else:      contagem[nome]["Diurno FimSemana"] += 1
            break
    # Noturno
    for j in range(len(plantonistas)):
        nome = plantonistas[(start_rr+j) % len(plantonistas)]
        if (data,"noturno") in ocorrencias_por_nome[nome]:
            if not in_ferias(nome, data, ferias_plant):
                ws_out.cell(row=row, column=COL_NOTURNO, value=nome)
                dw = data.weekday()
                if dw in (0,1,2,3): contagem[nome]["Noturno Semana"] += 1
                else:               contagem[nome]["Noturno FimSemana"] += 1
            break
    start_rr = (start_rr+1) % len(plantonistas)

# Etapa 2: Expedientes nos noturnos de fim de semana (balanceado)
counts_exped = {nome: 0 for nome in expedientes}

def escolher_expediente_balanceado(data_atual):
    disponiveis = [n for n in expedientes if not in_ferias(n, data_atual, ferias_exped)]
    if not disponiveis:
        return None
    min_count = min(counts_exped[n] for n in disponiveis)
    candidatos = [n for n in disponiveis if counts_exped[n] == min_count]
    return candidatos[0] if candidatos else None

for row, data in datas:
    if data.weekday() in (4,5,6):  # sexta, sábado, domingo
        if not ws_out.cell(row=row, column=COL_NOTURNO).value:
            nome_exp = escolher_expediente_balanceado(data)
            if nome_exp:
                ws_out.cell(row=row, column=COL_NOTURNO, value=nome_exp)
                counts_exped[nome_exp] += 1
                contagem[nome_exp]["Noturno FimSemana"] += 1

# Etapa 3: Expedientes nos noturnos de semana (seg-qui), balanceando totais
for row, data in datas:
    if data.weekday() in (0,1,2,3):  # seg-qui
        if not ws_out.cell(row=row, column=COL_NOTURNO).value:
            disponiveis = [n for n in expedientes if not in_ferias(n, data, ferias_exped)]
            if disponiveis:
                min_total = min(counts_exped[n] for n in disponiveis)
                candidatos = [n for n in disponiveis if counts_exped[n] == min_total]
                nome_exp = candidatos[0]
                ws_out.cell(row=row, column=COL_NOTURNO, value=nome_exp)
                counts_exped[nome_exp] += 1
                contagem[nome_exp]["Noturno Semana"] += 1

# Aba Resumo
ws_resumo = wb_out.create_sheet("Resumo")
ws_resumo.append(["Delegado","Diurno Semana","Noturno Semana","Diurno FimSemana","Noturno FimSemana","Total"])
for nome, dados in sorted(contagem.items()):
    total = sum(dados.values())
    ws_resumo.append([nome,
                      dados["Diurno Semana"],
                      dados["Notorno Semana"] if "Notorno Semana" in dados else dados["Noturno Semana"],
                      dados["Diurno FimSemana"],
                      dados["Notorno FimSemana"] if "Notorno FimSemana" in dados else dados["Noturno FimSemana"],
                      total])

# Largura das colunas do resumo
for col in ["A","B","C","D","E","F"]:
    ws_resumo.column_dimensions[col].width = 25

wb_out.save(saida_final)
print(f"✅ Escala final salva em: {saida_final}")
