import pandas as pd
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime, timedelta

# -------------------------
# Utilitários
# -------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        try:
            return pd.to_datetime(cell).date()
        except:
            return None

def in_ferias(nome, data, ferias_map):
    return any(i <= data <= f for i, f in ferias_map.get(nome, []))

# -------------------------
# Configuração de arquivos
# -------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados2.xlsx"
arquivo_base      = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final       = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL.xlsx"

# -------------------------
# Carregar delegados e ferias
# -------------------------
df_delegados = pd.read_excel(arquivo_delegados)

df_plantonistas = df_delegados[df_delegados.iloc[:,1].astype(str).str.lower().str.contains("plantão")]
plantonistas = df_plantonistas.iloc[:,0].dropna().astype(str).str.strip().tolist()

df_expedientes = df_delegados[df_delegados.iloc[:,1].astype(str).str.lower().str.contains("expediente")]
expedientes = df_expedientes.iloc[:,0].dropna().astype(str).str.strip().tolist()

def map_ferias(df):
    ferias = {}
    for _, row in df.iterrows():
        nome = str(row.iloc[0]).strip()
        periodos = []
        for i in [1, 2]:
            ini_col = f"Inicio Férias {i}"
            fim_col = f"Término Férias {i}"
            if ini_col in df.columns and fim_col in df.columns:
                ini = row.get(ini_col)
                fim = row.get(fim_col)
                if pd.notna(ini) and pd.notna(fim):
                    try:
                        periodos.append((pd.to_datetime(ini).date(), pd.to_datetime(fim).date()))
                    except:
                        pass
        if periodos:
            ferias[nome] = periodos
    return ferias

ferias_plant = map_ferias(df_plantonistas)
ferias_exped = map_ferias(df_expedientes)

# -------------------------
# Carregar planilha base e copiar layout (leitura única)
# -------------------------
wb_base = load_workbook(arquivo_base)
ws_base = wb_base.active
COL_DATA, COL_DIURNO, COL_NOTURNO = 1, 3, 4

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Escala"

# copiar layout e estilos uma única vez (mantendo células)
for row in ws_base.iter_rows():
    for cell in row:
        new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)

for col_letter, dim in ws_base.column_dimensions.items():
    try:
        ws_out.column_dimensions[col_letter].width = dim.width
    except:
        pass
for row_num, dim in ws_base.row_dimensions.items():
    try:
        ws_out.row_dimensions[row_num].height = dim.height
    except:
        pass
for merged_range in ws_base.merged_cells.ranges:
    ws_out.merge_cells(str(merged_range))
ws_out.freeze_panes = ws_base.freeze_panes

# -------------------------
# Ler as datas e construir estrutura em memória (lista + dict)
# -------------------------
datas = []
# schedule_map: row -> {'date': date, 'diurno': value, 'noturno': value}
schedule_map = {}

max_row = ws_base.max_row
for r in range(2, max_row+1):
    d = parse_data(ws_base.cell(row=r, column=COL_DATA).value)
    if d:
        datas.append((r, d))
        # armazenar valores iniciais (pode estar vazio)
        schedule_map[r] = {
            'date': d,
            'diurno': ws_out.cell(row=r, column=COL_DIURNO).value,
            'noturno': ws_out.cell(row=r, column=COL_NOTURNO).value
        }

if not datas:
    raise SystemExit("Nenhuma data válida encontrada na planilha base.")

first_date = datas[0][1]
last_date  = datas[-1][1]

# -------------------------
# Plantonistas: gerar ocorrencias (ciclo 12x24-12x72)
# -------------------------
ocorrencias_por_nome = {nome: set() for nome in plantonistas}
for i, nome in enumerate(plantonistas):
    offset_days = i % 5
    current_date = first_date + timedelta(days=offset_days)
    turno_diurno = True
    while current_date <= last_date:
        period = "diurno" if turno_diurno else "noturno"
        ocorrencias_por_nome[nome].add((current_date, period))
        step_days = 1 if turno_diurno else 4
        current_date += timedelta(days=step_days)
        turno_diurno = not turno_diurno

# -------------------------
# Contagem e contadores iniciais
# -------------------------
contagem = {nome: {"Diurno Semana":0,"Noturno Semana":0,"Diurno FimSemana":0,"Noturno FimSemana":0}
            for nome in plantonistas+expedientes}
counts_exped = {nome: 0 for nome in expedientes}

# -------------------------
# Helpers de agenda em memória (rápidos)
# -------------------------
def get_slot(row, col):
    entry = schedule_map.get(row)
    if not entry:
        return None
    return entry['diurno'] if col == COL_DIURNO else entry['noturno']

def set_slot(row, col, nome):
    if row not in schedule_map:
        schedule_map[row] = {'date': None, 'diurno': None, 'noturno': None}
    if col == COL_DIURNO:
        schedule_map[row]['diurno'] = nome
    else:
        schedule_map[row]['noturno'] = nome

def aparece_na_linha_mem(row, nome):
    if row not in schedule_map:
        return False
    s = schedule_map[row]
    return s['diurno'] == nome or s['noturno'] == nome

# -------------------------
# Função de alocação aprimorada (usa apenas a memória)
# inclui shifts, trocas lógicas e preenchimento forçado
# -------------------------
def tentar_alocar_com_troca_mem(nome, idx_atual, datas, col, contagem_dict, counts_dict, dias_validos, ferias_exped):
    """
    Versão aprimorada em memória que:
      - tenta posição atual
      - tenta frente
      - tenta trás
      - troca lógica (somente entre expedientes)
      - shift de blocos para frente/back (cadeias de expedientes)
      - preenchimento forçado como último recurso
    """

    def pode_alocar_sim(nome, row, sched):
        """Verifica se 'nome' poderia ser alocado em 'row' no schedule 'sched' (dicionário)."""
        if row not in sched:
            return False
        s = sched[row]
        # não pode estar já nessa linha
        if s['diurno'] == nome or s['noturno'] == nome:
            return False
        # anterior
        prev = sched.get(row-1)
        if prev and (prev['diurno'] == nome or prev['noturno'] == nome):
            return False
        # seguinte
        nxt = sched.get(row+1)
        if nxt and (nxt['diurno'] == nome or nxt['noturno'] == nome):
            return False
        return True

    def registrar_local(nome, row, data):
        """Atualiza slot e contadores (usado ao confirmar alocação)."""
        set_slot(row, col, nome)
        if nome in counts_dict:
            counts_dict[nome] += 1
        if nome in contagem_dict:
            wd = data.weekday()
            if col == COL_NOTURNO:
                if wd < 5:
                    contagem_dict[nome]["Noturno Semana"] += 1
                else:
                    contagem_dict[nome]["Noturno FimSemana"] += 1
            else:
                if wd < 5:
                    contagem_dict[nome]["Diurno Semana"] += 1
                else:
                    contagem_dict[nome]["Diurno FimSemana"] += 1

    row_atual, data_atual = datas[idx_atual]

    # 1) posição atual
    if get_slot(row_atual, col) in (None, ""):
        if not in_ferias(nome, data_atual, ferias_exped) and pode_alocar_sim(nome, row_atual, schedule_map):
            registrar_local(nome, row_atual, data_atual)
            return True

    # 2) frente
    for fut_idx in range(idx_atual+1, len(datas)):
        fut_row, fut_data = datas[fut_idx]
        if fut_data.weekday() not in dias_validos:
            continue
        if get_slot(fut_row, col) not in (None, ""):
            continue
        if in_ferias(nome, fut_data, ferias_exped):
            continue
        if not pode_alocar_sim(nome, fut_row, schedule_map):
            continue
        registrar_local(nome, fut_row, fut_data)
        return True

    # 3) trás (sem troca)
    for prev_idx in range(idx_atual-1, -1, -1):
        prev_row, prev_data = datas[prev_idx]
        if prev_data.weekday() not in dias_validos:
            continue
        if get_slot(prev_row, col) not in (None, ""):
            continue
        if in_ferias(nome, prev_data, ferias_exped):
            continue
        if not pode_alocar_sim(nome, prev_row, schedule_map):
            continue
        registrar_local(nome, prev_row, prev_data)
        return True

    # 4) troca lógica (somente entre expedientes)
    for prev_idx in range(idx_atual-1, -1, -1):
        prev_row, prev_data = datas[prev_idx]
        if prev_data.weekday() not in dias_validos:
            continue
        ocupante = get_slot(prev_row, col)
        if not ocupante:
            continue
        # impedir troca com plantonista
        if ocupante not in counts_dict:
            continue
        # ocupante não pode estar de ferias no dia da troca
        if in_ferias(ocupante, data_atual, ferias_exped):
            continue
        # ambos devem respeitar proximidade nas novas posições
        if not pode_alocar_sim(nome, prev_row, schedule_map):
            continue
        if not pode_alocar_sim(ocupante, row_atual, schedule_map):
            continue

        # efetua troca em memória
        set_slot(prev_row, col, nome)
        set_slot(row_atual, col, ocupante)

        # ajustar contagens
        if nome in counts_dict:
            counts_dict[nome] += 1
        if ocupante in counts_dict:
            counts_dict[ocupante] += 1

        # ajustar contagem por dia/turno
        wd_prev = prev_data.weekday()
        if nome in contagem_dict:
            if col == COL_NOTURNO:
                if wd_prev < 5:
                    contagem_dict[nome]["Noturno Semana"] += 1
                else:
                    contagem_dict[nome]["Noturno FimSemana"] += 1
            else:
                if wd_prev < 5:
                    contagem_dict[nome]["Diurno Semana"] += 1
                else:
                    contagem_dict[nome]["Diurno FimSemana"] += 1

        wd_atual = data_atual.weekday()
        if ocupante in contagem_dict:
            if col == COL_NOTURNO:
                if wd_atual < 5:
                    contagem_dict[ocupante]["Noturno Semana"] += 1
                else:
                    contagem_dict[ocupante]["Noturno FimSemana"] += 1
            else:
                if wd_atual < 5:
                    contagem_dict[ocupante]["Diurno Semana"] += 1
                else:
                    contagem_dict[ocupante]["Diurno FimSemana"] += 1

        return True

    # 4b) SHIFT DE BLOCOS PARA FRENTE (empurrar cadeia para frente)
    MAX_SHIFT = 8  # ajustável
    for depth in range(1, MAX_SHIFT+1):
        start_row = row_atual - depth
        if start_row < 2:
            break
        chain_rows = list(range(start_row, row_atual))
        ok_chain = True
        for rr in chain_rows:
            occ = get_slot(rr, col)
            if not occ or occ not in counts_dict:  # vazio ou plantonista -> não serve
                ok_chain = False
                break
        if not ok_chain:
            continue

        # simulate moves on a temp schedule copy
        sched_copy = {r: {'diurno': v['diurno'], 'noturno': v['noturno'], 'date': v['date']} for r, v in schedule_map.items()}

        can_shift = True
        for rr in reversed(chain_rows):
            src_name = sched_copy[rr]['diurno'] if col == COL_DIURNO else sched_copy[rr]['noturno']
            dest_row = rr + 1
            dest_date = sched_copy[dest_row]['date']
            if in_ferias(src_name, dest_date, ferias_exped):
                can_shift = False
                break
            # vacate src
            if col == COL_DIURNO:
                sched_copy[rr]['diurno'] = None
            else:
                sched_copy[rr]['noturno'] = None
            # adjacency checks at dest_row
            dest_prev = sched_copy.get(dest_row-1)
            dest_next = sched_copy.get(dest_row+1)
            if ((dest_prev and (dest_prev['diurno'] == src_name or dest_prev['noturno'] == src_name)) or
                (dest_next and (dest_next['diurno'] == src_name or dest_next['noturno'] == src_name)) or
                (sched_copy[dest_row]['diurno'] == src_name if col == COL_DIURNO else sched_copy[dest_row]['noturno'] == src_name)):
                can_shift = False
                break
            # perform move on simulated schedule
            if col == COL_DIURNO:
                sched_copy[dest_row]['diurno'] = src_name
            else:
                sched_copy[dest_row]['noturno'] = src_name

        if can_shift:
            # commit simulated shifts to real schedule_map and update counts/contagem
            for rr in reversed(chain_rows):
                src_name = schedule_map[rr]['diurno'] if col == COL_DIURNO else schedule_map[rr]['noturno']
                dest_row = rr + 1
                dest_date = schedule_map[dest_row]['date']
                # move
                if col == COL_DIURNO:
                    schedule_map[rr]['diurno'] = None
                    schedule_map[dest_row]['diurno'] = src_name
                else:
                    schedule_map[rr]['noturno'] = None
                    schedule_map[dest_row]['noturno'] = src_name
                # update counters
                if src_name in counts_dict:
                    counts_dict[src_name] += 1
                if src_name in contagem_dict:
                    if col == COL_NOTURNO:
                        if dest_date.weekday() < 5:
                            contagem_dict[src_name]["Noturno Semana"] += 1
                        else:
                            contagem_dict[src_name]["Noturno FimSemana"] += 1
                    else:
                        if dest_date.weekday() < 5:
                            contagem_dict[src_name]["Diurno Semana"] += 1
                        else:
                            contagem_dict[src_name]["Diurno FimSemana"] += 1
            return True

    # 4c) SHIFT DE BLOCOS PARA TRÁS (puxar cadeia para trás)
    for depth in range(1, MAX_SHIFT+1):
        end_row = row_atual + depth
        if end_row > max_row:
            break
        chain_rows = list(range(row_atual+1, end_row+1))
        ok_chain = True
        for rr in chain_rows:
            occ = get_slot(rr, col)
            if not occ or occ not in counts_dict:
                ok_chain = False
                break
        if not ok_chain:
            continue

        sched_copy = {r: {'diurno': v['diurno'], 'noturno': v['noturno'], 'date': v['date']} for r, v in schedule_map.items()}
        can_shift = True
        for rr in chain_rows:
            src_name = sched_copy[rr]['diurno'] if col == COL_DIURNO else sched_copy[rr]['noturno']
            dest_row = rr - 1
            dest_date = sched_copy[dest_row]['date']
            if in_ferias(src_name, dest_date, ferias_exped):
                can_shift = False
                break
            if col == COL_DIURNO:
                sched_copy[rr]['diurno'] = None
            else:
                sched_copy[rr]['noturno'] = None
            dest_prev = sched_copy.get(dest_row-1)
            dest_next = sched_copy.get(dest_row+1)
            if ((dest_prev and (dest_prev['diurno'] == src_name or dest_prev['noturno'] == src_name)) or
                (dest_next and (dest_next['diurno'] == src_name or dest_next['noturno'] == src_name)) or
                (sched_copy[dest_row]['diurno'] == src_name if col == COL_DIURNO else sched_copy[dest_row]['noturno'] == src_name)):
                can_shift = False
                break
            if col == COL_DIURNO:
                sched_copy[dest_row]['diurno'] = src_name
            else:
                sched_copy[dest_row]['noturno'] = src_name

        if can_shift:
            # commit simulated pulls
            for rr in chain_rows:
                src_name = schedule_map[rr]['diurno'] if col == COL_DIURNO else schedule_map[rr]['noturno']
                dest_row = rr - 1
                dest_date = schedule_map[dest_row]['date']
                if col == COL_DIURNO:
                    schedule_map[rr]['diurno'] = None
                    schedule_map[dest_row]['diurno'] = src_name
                else:
                    schedule_map[rr]['noturno'] = None
                    schedule_map[dest_row]['noturno'] = src_name
                if src_name in counts_dict:
                    counts_dict[src_name] += 1
                if src_name in contagem_dict:
                    if col == COL_NOTURNO:
                        if dest_date.weekday() < 5:
                            contagem_dict[src_name]["Noturno Semana"] += 1
                        else:
                            contagem_dict[src_name]["Noturno FimSemana"] += 1
                    else:
                        if dest_date.weekday() < 5:
                            contagem_dict[src_name]["Diurno Semana"] += 1
                        else:
                            contagem_dict[src_name]["Diurno FimSemana"] += 1
            return True

    # 5) preenchimento forçado (último recurso) — escolher expediente disponível com menor carga
    candidatos = [n for n in counts_dict.keys() if not in_ferias(n, data_atual, ferias_exped)]
    if candidatos:
        escolhido = min(candidatos, key=lambda n: counts_dict[n])
        set_slot(row_atual, col, escolhido)
        counts_dict[escolhido] += 1
        if escolhido in contagem_dict:
            wd = data_atual.weekday()
            if col == COL_NOTURNO:
                if wd < 5:
                    contagem_dict[escolhido]["Noturno Semana"] += 1
                else:
                    contagem_dict[escolhido]["Noturno FimSemana"] += 1
            else:
                if wd < 5:
                    contagem_dict[escolhido]["Diurno Semana"] += 1
                else:
                    contagem_dict[escolhido]["Diurno FimSemana"] += 1
        return True

    return False

# -------------------------
# ETAPA 1: Preencher plantonistas (usa schedule_map)
# -------------------------
start_rr = 0
for idx, (row, data) in enumerate(datas):
    # Diurno
    for j in range(len(plantonistas)):
        nome = plantonistas[(start_rr+j) % len(plantonistas)]
        if (data,"diurno") in ocorrencias_por_nome[nome]:
            if not in_ferias(nome, data, ferias_plant):
                set_slot(row, COL_DIURNO, nome)
                dw = data.weekday()
                if dw < 5:
                    contagem[nome]["Diurno Semana"] += 1
                else:
                    contagem[nome]["Diurno FimSemana"] += 1
            break

    # Noturno
    for j in range(len(plantonistas)):
        nome = plantonistas[(start_rr+j) % len(plantonistas)]
        if (data,"noturno") in ocorrencias_por_nome[nome]:
            if not in_ferias(nome, data, ferias_plant):
                set_slot(row, COL_NOTURNO, nome)
                dw = data.weekday()
                if dw in (0,1,2,3):
                    contagem[nome]["Noturno Semana"] += 1
                else:
                    contagem[nome]["Noturno FimSemana"] += 1
            break

    start_rr = (start_rr+1) % len(plantonistas)

# -------------------------
# ETAPA 2: Expedientes noturno fim de semana (balanceado) — usa memória
# -------------------------
def escolher_expediente_balanceado_mem(data_atual):
    disponiveis = [n for n in expedientes if not in_ferias(n, data_atual, ferias_exped)]
    if not disponiveis:
        return None
    min_count = min(counts_exped[n] for n in disponiveis)
    candidatos = [n for n in disponiveis if counts_exped[n] == min_count]
    return candidatos[0] if candidatos else None

for row, data in datas:
    if data.weekday() in (4,5,6):  # sexta, sab, dom
        if get_slot(row, COL_NOTURNO) in (None, ""):
            nome_exp = escolher_expediente_balanceado_mem(data)
            if nome_exp:
                set_slot(row, COL_NOTURNO, nome_exp)
                counts_exped[nome_exp] += 1
                contagem[nome_exp]["Noturno FimSemana"] += 1

# -------------------------
# ETAPA 3: Noturnos de semana com realocação/troca
# -------------------------
for idx, (row, data) in enumerate(datas):
    if data.weekday() in (0,1,2,3):
        if get_slot(row, COL_NOTURNO) in (None, ""):
            disponiveis = [n for n in expedientes if not in_ferias(n, data, ferias_exped)]
            if disponiveis:
                min_total = min(counts_exped[n] for n in disponiveis)
                candidatos = [n for n in disponiveis if counts_exped[n] == min_total]
                for nome in candidatos:
                    if tentar_alocar_com_troca_mem(nome, idx, datas, COL_NOTURNO, contagem, counts_exped, (0,1,2,3), ferias_exped):
                        break

# -------------------------
# ETAPA 4A: Diurnos fim de semana (com troca)
# -------------------------
for idx, (row, data) in enumerate(datas):
    if data.weekday() in (5,6):
        if get_slot(row, COL_DIURNO) in (None, ""):
            disponiveis = [n for n in expedientes if not in_ferias(n, data, ferias_exped)]
            if disponiveis:
                min_total = min(counts_exped[n] for n in disponiveis)
                candidatos = [n for n in disponiveis if counts_exped[n] == min_total]
                for nome in candidatos:
                    if tentar_alocar_com_troca_mem(nome, idx, datas, COL_DIURNO, contagem, counts_exped, (5,6), ferias_exped):
                        break

# -------------------------
# ETAPA 4B: Diurnos semana (com troca)
# -------------------------
for idx, (row, data) in enumerate(datas):
    if data.weekday() in (0,1,2,3,4):
        if get_slot(row, COL_DIURNO) in (None, ""):
            disponiveis = [n for n in expedientes if not in_ferias(n, data, ferias_exped)]
            if disponiveis:
                min_total = min(counts_exped[n] for n in disponiveis)
                candidatos = [n for n in disponiveis if counts_exped[n] == min_total]
                for nome in candidatos:
                    if tentar_alocar_com_troca_mem(nome, idx, datas, COL_DIURNO, contagem, counts_exped, (0,1,2,3,4), ferias_exped):
                        break

# -------------------------
# Módulo final: Correção global (garante zero lacunas)
# percorre a tabela e tenta preencher quaisquer lacunas restantes
# usando a mesma lógica (chamando a função de alocação em memória).
# -------------------------
def corrigir_lacunas_global(max_passes=8):
    changed = True
    passes = 0
    while changed and passes < max_passes:
        changed = False
        passes += 1
        # percorre todas as datas; tenta preencher diurno e noturno
        for idx, (row, data) in enumerate(datas):
            # DIURNO
            if get_slot(row, COL_DIURNO) in (None, ""):
                # candidatos ordenados por menor carga
                candidatos = sorted([n for n in expedientes if not in_ferias(n, data, ferias_exped)], key=lambda n: counts_exped[n])
                for nome in candidatos:
                    dias_validos = (5,6) if data.weekday() in (5,6) else (0,1,2,3,4)
                    if tentar_alocar_com_troca_mem(nome, idx, datas, COL_DIURNO, contagem, counts_exped, dias_validos, ferias_exped):
                        changed = True
                        break
            # NOTURNO
            if get_slot(row, COL_NOTURNO) in (None, ""):
                candidatos = sorted([n for n in expedientes if not in_ferias(n, data, ferias_exped)], key=lambda n: counts_exped[n])
                # definir dias_validos para noturno: se sexta/sab/dom usar (4,5,6) conforme sua regra, senão (0,1,2,3)
                if data.weekday() in (4,5,6):
                    noturno_valid = (4,5,6)
                else:
                    noturno_valid = (0,1,2,3)
                for nome in candidatos:
                    if tentar_alocar_com_troca_mem(nome, idx, datas, COL_NOTURNO, contagem, counts_exped, noturno_valid, ferias_exped):
                        changed = True
                        break
    # fim while
    return

corrigir_lacunas_global(max_passes=12)

# -------------------------
# Passo final: escrevendo de uma vez no Excel (uma única passagem de I/O)
# -------------------------
for row, entry in schedule_map.items():
    # só escrever se existir algo (preserva estilos copiados)
    if 'diurno' in entry:
        ws_out.cell(row=row, column=COL_DIURNO, value=entry['diurno'])
    if 'noturno' in entry:
        ws_out.cell(row=row, column=COL_NOTURNO, value=entry['noturno'])

# -------------------------
# Aba Resumo
# -------------------------
ws_resumo = wb_out.create_sheet("Resumo")
ws_resumo.append([
    "Delegado",
    "Diurno Semana",
    "Noturno Semana",
    "Diurno FimSemana",
    "Noturno FimSemana",
    "Total"
])

for nome, dados in sorted(contagem.items()):
    total = (dados["Diurno Semana"] + dados["Noturno Semana"] +
             dados["Diurno FimSemana"] + dados["Noturno FimSemana"])
    ws_resumo.append([
        nome,
        dados["Diurno Semana"],
        dados["Noturno Semana"],
        dados["Diurno FimSemana"],
        dados["Noturno FimSemana"],
        total
    ])

for col in ["A","B","C","D","E","F"]:
    ws_resumo.column_dimensions[col].width = 25

wb_out.save(saida_final)
print(f"✅ Escala final salva em: {saida_final}")
