import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import deque

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Arquivos
# -----------------------------
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_PLANTONISTAS.xlsx"

# -----------------------------
# Configurações dos plantonistas
# -----------------------------
plantonistas = ["A", "B", "C"]  # nomes dos plantonistas

# Exemplo de férias (opcional)
ferias = {
    # "A": [(datetime(2025,1,10).date(), datetime(2025,1,15).date())],
}

# -----------------------------
# Carregar planilha existente
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active

col_data = 1     # coluna A
col_diurno = 3   # coluna C
col_noturno = 4  # coluna D

# -----------------------------
# Inicializar fila cíclica dos plantonistas
# -----------------------------
fila = deque(plantonistas)
estado_plantao = {p:0 for p in plantonistas}  # folga restante em turnos
proximo_ciclo = {p:0 for p in plantonistas}   # 0 = 12x24, 1 = 12x72

# -----------------------------
# Preencher Noturno Semana (segunda a quinta) com ciclo 12x24 + 12x72
# -----------------------------
for row in range(2, ws.max_row + 1):
    data = parse_data(ws.cell(row=row, column=col_data).value)
    if not data: 
        continue

    dia_semana = data.weekday()  # 0=segunda ... 6=domingo
    if dia_semana not in (0,1,2,3):  # segunda a quinta
        continue

    cell_noturno = ws.cell(row=row, column=col_noturno)
    if cell_noturno.value not in (None,""," "):
        continue  # já preenchido

    # Percorre a fila até achar plantonista disponível
    for _ in range(len(fila)):
        p = fila[0]
        # verifica férias
        if p in ferias and any(start <= data <= end for start,end in ferias[p]):
            fila.rotate(-1)
            continue
        # verifica folga
        if estado_plantao[p] > 0:
            fila.rotate(-1)
            continue

        # plantonista escolhido
        cell_noturno.value = p

        # aplica ciclo
        if proximo_ciclo[p] == 0:
            estado_plantao[p] = 2  # folga 24h = 2 turnos
            proximo_ciclo[p] = 1
        else:
            estado_plantao[p] = 6  # folga 72h = 6 turnos
            proximo_ciclo[p] = 0

        # move plantonista para o final da fila
        fila.rotate(-1)
        break

    # diminui contagem de folga de todos
    for p in estado_plantao:
        if estado_plantao[p] > 0:
            estado_plantao[p] -= 1

# -----------------------------
# Salvar planilha final
# -----------------------------
wb.save(saida_final)
print(f"✅ Planilha preenchida com plantonistas salva em: {saida_final}")
