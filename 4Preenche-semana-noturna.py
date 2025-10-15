import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Arquivos
# -----------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_COMPLETA.xlsx"

# -----------------------------
# Carregar delegados e férias
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()
mapa_codigos = dict(zip(df_delegados["Código"], df_delegados["Nome"]))

ferias = {}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    for i in [1, 2]:
        inicio = row.get(f"Inicio Férias {i}")
        fim = row.get(f"Término Férias {i}")
        if pd.notna(inicio) and pd.notna(fim):
            periodos.append((pd.to_datetime(inicio).date(), pd.to_datetime(fim).date()))
    if periodos:
        ferias[codigo] = periodos

# Delegados de rodízio (códigos 3–22)
delegados_rodizio = df_delegados[df_delegados["Código"].astype(int).between(3, 22)]
lista_delegados = list(zip(delegados_rodizio["Código"], delegados_rodizio["Nome"]))

# -----------------------------
# Carregar planilha da escala
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active
col_data, col_diurno, col_noturno = 1, 3, 4

# -----------------------------
# Padrão inicial de preenchimento
# -----------------------------
padrao = [("1","2"),("","1"),("2",""),("",""),("","")]
indice_padrao = 0

# -----------------------------
# Histórico usado para verificar folgas e mesmo dia
# -----------------------------
escala_historico = []

# -----------------------------
# Contagem de plantões
# -----------------------------
contagem_total = {nome: {"Diurno Semana":0,"Noturno Semana":0,"Diurno FimSemana":0,"Noturno FimSemana":0}
                  for nome in df_delegados["Nome"]}

contagem_fs_diurno = {codigo:0 for codigo,_ in lista_delegados}
contagem_fs_noturno = {codigo:0 for codigo,_ in lista_delegados}

# -----------------------------
# Funções auxiliares
# -----------------------------
def em_ferias(codigo, data):
    return any(inicio <= data <= fim for inicio, fim in ferias.get(codigo, []))

def tem_folga(nome, data):
    for _, d, c, n in escala_historico:
        if n == nome or c == nome:
            if abs((d - data).days) <= 1:
                return False
    return True

def trabalha_mesmo_dia(nome, data):
    for _, d, c, n in escala_historico:
        if d == data and (c == nome or n == nome):
            return True
    return False

def candidatos_equilibrados(contagem_dict):
    if not contagem_dict: return []
    minimo = min(contagem_dict.values())
    maximo = minimo + 1
    return [(c, n) for c, n in contagem_dict.items() if n <= maximo]

# -----------------------------
# Etapa 1: Preencher padrão inicial
# -----------------------------
for row in range(2, ws.max_row+1):
    data = parse_data(ws.cell(row=row, column=col_data).value)
    if not data: continue

    c_val,d_val = padrao[indice_padrao]
    indice_padrao = (indice_padrao + 1) % len(padrao)
    c_nome = mapa_codigos.get(c_val,"") if c_val else ""
    d_nome = mapa_codigos.get(d_val,"") if d_val else ""

    if c_val and em_ferias(c_val, data): c_nome = ""
    if d_val and em_ferias(d_val, data): d_nome = ""

    ws.cell(row=row, column=col_diurno, value=c_nome)
    ws.cell(row=row, column=col_noturno, value=d_nome)

    dia_semana = data.weekday()
    if dia_semana < 4:  # segunda a quinta
        if c_nome: contagem_total[c_nome]["Diurno Semana"] +=1
        if d_nome: contagem_total[d_nome]["Noturno Semana"] +=1
    elif dia_semana == 4:  # sexta
        if c_nome: contagem_total[c_nome]["Diurno Semana"] +=1
        if d_nome: contagem_total[d_nome]["Noturno FimSemana"] +=1
    elif dia_semana in (5,6):  # sábado e domingo
        if c_nome: contagem_total[c_nome]["Diurno FimSemana"] +=1
        if d_nome: contagem_total[d_nome]["Noturno FimSemana"] +=1

    escala_historico.append((row,data,c_nome,d_nome))

# -----------------------------
# Etapa 2: Preencher lacunas FDS (diurno e noturno)
# -----------------------------
for row in range(2, ws.max_row+1):
    data = parse_data(ws.cell(row=row, column=col_data).value)
    if not data: continue
    dia_semana = data.weekday()

    # Diurno FDS (sábado e domingo)
    cell_diurno = ws.cell(row=row, column=col_diurno)
    if dia_semana in (5,6) and cell_diurno.value in (None,""," "):
        candidatos = sorted(candidatos_equilibrados(contagem_fs_diurno), key=lambda x:x[1])
        for codigo,_ in candidatos:
            nome = mapa_codigos[codigo]
            if em_ferias(codigo,data): continue
            if trabalha_mesmo_dia(nome,data): continue
            if not tem_folga(nome,data): continue
            cell_diurno.value = nome
            contagem_fs_diurno[codigo] +=1
            contagem_total[nome]["Diurno FimSemana"] +=1
            escala_historico.append((row,data,nome,ws.cell(row=row,column=col_noturno).value))
            break

    # Noturno FDS (sexta, sábado, domingo)
    cell_noturno = ws.cell(row=row, column=col_noturno)
    if dia_semana in (4,5,6) and cell_noturno.value in (None,""," "):
        candidatos = sorted(candidatos_equilibrados(contagem_fs_noturno), key=lambda x:x[1])
        for codigo,_ in candidatos:
            nome = mapa_codigos[codigo]
            if em_ferias(codigo,data): continue
            if trabalha_mesmo_dia(nome,data): continue
            if not tem_folga(nome,data): continue
            cell_noturno.value = nome
            contagem_fs_noturno[codigo] +=1
            contagem_total[nome]["Noturno FimSemana"] +=1
            escala_historico.append((row,data,ws.cell(row=row,column=col_diurno).value,nome))
            break

# -----------------------------
# Etapa 3: Preencher lacunas Noturno Semana (segunda a quinta)
# -----------------------------
for row in range(2, ws.max_row + 1):
    data = parse_data(ws.cell(row=row, column=col_data).value)
    if not data: continue
    dia_semana = data.weekday()
    if dia_semana in (0,1,2,3):
        cell_noturno = ws.cell(row=row, column=col_noturno)
        if cell_noturno.value in (None,""," "):
            candidatos = sorted(candidatos_equilibrados(contagem_fs_noturno), key=lambda x:x[1])
            for codigo,_ in candidatos:
                nome = mapa_codigos[codigo]
                if em_ferias(codigo,data): continue
                if trabalha_mesmo_dia(nome,data): continue
                if not tem_folga(nome,data): continue
                cell_noturno.value = nome
                contagem_fs_noturno[codigo] +=1
                contagem_total[nome]["Noturno Semana"] +=1
                escala_historico.append((row,data,ws.cell(row=row,column=col_diurno).value,nome))
                break

# -----------------------------
# Gera aba de resumo final
# -----------------------------
if "Resumo" in wb.sheetnames:
    del wb["Resumo"]
ws_resumo = wb.create_sheet("Resumo")
ws_resumo.append(["Delegado","Diurno Semana","Noturno Semana","Diurno FimSemana","Noturno FimSemana","Total"])

for nome,dados in sorted(contagem_total.items()):
    total = sum(dados.values())
    ws_resumo.append([nome,dados["Diurno Semana"],dados["Noturno Semana"],
                      dados["Diurno FimSemana"],dados["Noturno FimSemana"],total])

for col in ["A","B","C","D","E","F"]:
    ws_resumo.column_dimensions[col].width = 25

# -----------------------------
# Salvar resultado final
# -----------------------------
wb.save(saida_final)
print(f"✅ Escala final completa salva em: {saida_final}")
