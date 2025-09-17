import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Caminhos dos arquivos
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025 plantonistas.xlsx"
saida_corrigida = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_CORRIGIDA_FORMATADA.xlsx"

# --- Lê delegados e férias ---
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str)  # força código como string

for col in ["Inicio Férias 1", "Término Férias 1", "Inicio Férias 2", "Término Férias 2"]:
    df_delegados[col] = pd.to_datetime(df_delegados[col], errors="coerce")

# --- Abre a planilha da escala mantendo formatação ---
wb = load_workbook(arquivo_escala)
ws = wb.active  # assume que a primeira aba é a escala

# Ajuste: qual coluna é data, diurno e noturno
col_data = 1         # coluna A (Data)
col_diurno = 3       # coluna C
col_noturno = 4      # coluna D

# --- Percorre cada linha da planilha ---
for row in range(2, ws.max_row + 1):  # assume que linha 1 é cabeçalho
    # Lê a data
    cell_data = ws.cell(row=row, column=col_data).value
    if isinstance(cell_data, datetime):
        data = cell_data
    else:
        try:
            data = pd.to_datetime(cell_data)
        except:
            continue

    for col in [col_diurno, col_noturno]:
        codigo = ws.cell(row=row, column=col).value
        if codigo is None:
            continue

        codigo_str = str(codigo).strip()  # garante string
        ferias = df_delegados[df_delegados["Código"] == codigo_str]

        em_ferias = False
        for _, f in ferias.iterrows():
            # Período 1
            if pd.notna(f["Inicio Férias 1"]) and pd.notna(f["Término Férias 1"]):
                if f["Inicio Férias 1"] <= data <= f["Término Férias 1"]:
                    em_ferias = True
            # Período 2
            if pd.notna(f["Inicio Férias 2"]) and pd.notna(f["Término Férias 2"]):
                if f["Inicio Férias 2"] <= data <= f["Término Férias 2"]:
                    em_ferias = True

        if em_ferias:
            # Apaga o valor da célula mantendo formatação
            ws.cell(row=row, column=col, value="")

# --- Salva a planilha mantendo formatação ---
wb.save(saida_corrigida)
print(f"✅ Escala corrigida salva em: {saida_corrigida}")
