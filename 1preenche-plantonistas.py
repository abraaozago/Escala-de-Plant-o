from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Carregar planilhas
# -----------------------------
delegados_wb = load_workbook("delegados.xlsx")
delegados_ws = delegados_wb.active

escala_wb = load_workbook("ESCALA 2º Semestre 2025.xlsx")
escala_ws = escala_wb.active

# -----------------------------
# Criar dicionário de férias por código
# -----------------------------
ferias = {}  # {codigo: [(inicio1, fim1), (inicio2, fim2)]}

for row in range(2, delegados_ws.max_row + 1):
    codigo = delegados_ws[f"B{row}"].value
    if not codigo:
        continue

    inicio1 = parse_data(delegados_ws[f"G{row}"].value)
    fim1 = parse_data(delegados_ws[f"H{row}"].value)
    inicio2 = parse_data(delegados_ws[f"I{row}"].value)
    fim2 = parse_data(delegados_ws[f"J{row}"].value)

    periodos = []
    if inicio1 and fim1:
        periodos.append((inicio1, fim1))
    if inicio2 and fim2:
        periodos.append((inicio2, fim2))
    
    ferias[codigo] = periodos

# -----------------------------
# Padrão de preenchimento
# -----------------------------
padrao = [("1", "2"), ("", "1"), ("2", ""), ("", ""), ("", "")]
indice = 0
linhas_preenchidas = []

# -----------------------------
# Preencher escala
# -----------------------------
# Supondo que a coluna A contém a data da escala
for row in range(2, escala_ws.max_row + 1):
    codigo = escala_ws[f"B{row}"].value
    data_celula = escala_ws[f"A{row}"].value  # coluna com a data da escala
    if not codigo or not data_celula:
        continue
    data = parse_data(data_celula)
    if not data:
        continue

    # Verificar se a data está em período de férias (+ dia anterior e posterior)
    em_ferias = False
    for inicio, fim in ferias.get(codigo, []):
        if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
            em_ferias = True
            break

    if not em_ferias:
        c_val, d_val = padrao[indice]
        escala_ws[f"C{row}"] = c_val
        escala_ws[f"D{row}"] = d_val
        indice = (indice + 1) % len(padrao)
        linhas_preenchidas.append((data, c_val, d_val))
    else:
        escala_ws[f"C{row}"] = ""
        escala_ws[f"D{row}"] = ""
        linhas_preenchidas.append((data, "", ""))

# -----------------------------
# Salvar planilha preenchida
# -----------------------------
escala_wb.save("ESCALA 2º Semestre 2025 plantonistas.xlsx")