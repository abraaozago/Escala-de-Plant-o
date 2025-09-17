import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Caminhos dos arquivos
# -----------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_NOMES.xlsx"

# -----------------------------
# Carregar delegados e férias
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()

# cria dicionário código → nome
mapa_codigos = dict(zip(df_delegados["Código"], df_delegados["Nome"]))

# converte colunas de férias
for col in ["Inicio Férias 1", "Término Férias 1", "Inicio Férias 2", "Término Férias 2"]:
    df_delegados[col] = pd.to_datetime(df_delegados[col], errors="coerce")

# monta dicionário de férias
ferias = {}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    if pd.notna(row["Inicio Férias 1"]) and pd.notna(row["Término Férias 1"]):
        periodos.append((row["Inicio Férias 1"].date(), row["Término Férias 1"].date()))
    if pd.notna(row["Inicio Férias 2"]) and pd.notna(row["Término Férias 2"]):
        periodos.append((row["Inicio Férias 2"].date(), row["Término Férias 2"].date()))
    if periodos:
        ferias[codigo] = periodos

# -----------------------------
# Carregar planilha da escala
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active

col_data = 1    # coluna A
col_diurno = 3  # coluna C
col_noturno = 4 # coluna D

# -----------------------------
# Padrão de preenchimento
# -----------------------------
# Aqui você pode ajustar os códigos dos delegados que entram no rodízio
padrao = [("1", "2"), ("", "1"), ("2", ""), ("", ""), ("", "")]
indice = 0

# -----------------------------
# Preencher escala
# -----------------------------
for row in range(2, ws.max_row + 1):
    cell_data = ws.cell(row=row, column=col_data).value
    if not cell_data:
        continue

    data = parse_data(cell_data)
    if not data:
        continue

    # pega os próximos códigos do padrão
    c_val, d_val = padrao[indice]
    indice = (indice + 1) % len(padrao)

    # substitui código por nome se existir no mapa
    c_nome = mapa_codigos.get(c_val, "") if c_val else ""
    d_nome = mapa_codigos.get(d_val, "") if d_val else ""

    # escreve provisoriamente
    ws.cell(row=row, column=col_diurno, value=c_nome)
    ws.cell(row=row, column=col_noturno, value=d_nome)

    # verifica férias do diurno
    if c_val and c_val in ferias:
        for inicio, fim in ferias[c_val]:
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_diurno, value="")

    # verifica férias do noturno
    if d_val and d_val in ferias:
        for inicio, fim in ferias[d_val]:
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_noturno, value="")

# -----------------------------
# Salvar planilha final
# -----------------------------
wb.save(saida_final)
print(f"✅ Escala final com nomes salva em: {saida_final}")
