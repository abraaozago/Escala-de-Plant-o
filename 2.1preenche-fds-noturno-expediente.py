import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# -----------------------------
# Função para converter célula em data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Arquivos
# -----------------------------
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_NOMES.xlsx"
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_NOMES_COMPLETA.xlsx"

# -----------------------------
# Carregar delegados e férias
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()

# Lista de delegados para rodízio (códigos 3–22)
delegados_rodizio = df_delegados[df_delegados["Código"].astype(int).between(3, 22)]
lista_delegados = list(zip(delegados_rodizio["Código"], delegados_rodizio["Nome"]))

# Cria dicionário de férias
ferias = {}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    for inicio_col, fim_col in [("Inicio Férias 1", "Término Férias 1"), ("Inicio Férias 2", "Término Férias 2")]:
        inicio = row[inicio_col]
        fim = row[fim_col]
        if pd.notna(inicio) and pd.notna(fim):
            periodos.append((inicio.date(), fim.date()))
    if periodos:
        ferias[codigo] = periodos

# -----------------------------
# Carregar planilha da escala
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active

col_data = 1    # coluna A
col_noturno = 4 # coluna D

# -----------------------------
# Preencher apenas finais de semana noturnos em branco
# -----------------------------
idx = 0
for row in range(2, ws.max_row + 1):
    cell_data = ws.cell(row=row, column=col_data).value
    data = parse_data(cell_data)
    if not data:
        continue

    # sexta (4), sábado (5) ou domingo (6)
    if data.weekday() in (4, 5, 6):
        cell_noturno = ws.cell(row=row, column=col_noturno)
        # só preencher se estiver em branco
        if cell_noturno.value in (None, "", " "):
            tentativas = 0
            while tentativas < len(lista_delegados):
                codigo, nome = lista_delegados[idx]
                # verifica se está em férias
                em_ferias = False
                if codigo in ferias:
                    for inicio, fim in ferias[codigo]:
                        if inicio <= data <= fim:
                            em_ferias = True
                            break
                if not em_ferias:
                    cell_noturno.value = nome
                    idx = (idx + 1) % len(lista_delegados)
                    break
                else:
                    idx = (idx + 1) % len(lista_delegados)
                    tentativas += 1
            # se todos estiverem de férias, a célula permanece em branco

# -----------------------------
# Salvar planilha final
# -----------------------------
wb.save(saida_final)
print(f"✅ Finais de semana noturnos preenchidos e escala salva em: {saida_final}")
