import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função auxiliar
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None


# -----------------------------
# ARQUIVOS
# -----------------------------
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_BALANCEADA_DETALHADA.xlsx"
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL_DIURNO_FDS_EQUILIBRADA.xlsx"

# -----------------------------
# CARREGA PLANILHA E DADOS
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active

df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str).str.strip()
mapa_codigos = dict(zip(df_delegados["Código"], df_delegados["Nome"]))

# Dicionário de férias
for col in ["Inicio Férias 1", "Término Férias 1", "Inicio Férias 2", "Término Férias 2"]:
    df_delegados[col] = pd.to_datetime(df_delegados[col], errors="coerce")

ferias = {}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    if pd.notna(row["Inicio Férias 1"]) and pd.notna(row["Término Férias 1"]):
        periodos.append((row["Inicio Férias 1"].date(), row["Término Férias 1"].date()))
    if pd.notna(row["Inicio Férias 2"]) and pd.notna(row["Término Férias 2"].date()):
        periodos.append((row["Inicio Férias 2"].date(), row["Término Férias 2"].date()))
    if periodos:
        ferias[codigo] = periodos

# Lista dos delegados de rodízio (3–22)
delegados_rodizio = df_delegados[df_delegados["Código"].astype(int).between(3, 22)]
lista_delegados = list(zip(delegados_rodizio["Código"], delegados_rodizio["Nome"]))

# -----------------------------
# LER DADOS EXISTENTES
# -----------------------------
col_data = 1
col_diurno = 3
col_noturno = 4

escala = []
for row in range(2, ws.max_row + 1):
    data = parse_data(ws.cell(row=row, column=col_data).value)
    if not data:
        continue
    diurno = ws.cell(row=row, column=col_diurno).value
    noturno = ws.cell(row=row, column=col_noturno).value
    escala.append((row, data, diurno, noturno))

# -----------------------------
# Identifica lacunas de diurno FDS (sábado e domingo)
# -----------------------------
lacunas_fds = [(r, d) for r, d, c, n in escala if d.weekday() in (5, 6) and (c in (None, "", " "))]

# Inicializa contagem dos plantões de fim de semana diurnos
contagem_fs_diurno = {codigo: 0 for codigo, _ in lista_delegados}

# Contabiliza plantões existentes
for _, _, diurno, _ in escala:
    if diurno:
        for codigo, nome in lista_delegados:
            if nome == diurno:
                contagem_fs_diurno[codigo] += 1

# -----------------------------
# Funções auxiliares
# -----------------------------
def em_ferias(codigo, data):
    if codigo not in ferias:
        return False
    for inicio, fim in ferias[codigo]:
        if inicio <= data <= fim:
            return True
    return False

def tem_folga(nome, data, escala):
    """Verifica se o delegado tem pelo menos 1 dia de folga antes e depois"""
    for _, d, c, n in escala:
        if n == nome or c == nome:
            if abs((d - data).days) <= 1:
                return False
    return True

def trabalha_mesmo_dia(nome, data, escala):
    """Verifica se o delegado já está escalado no mesmo dia (diurno ou noturno)"""
    for _, d, c, n in escala:
        if d == data and (c == nome or n == nome):
            return True
    return False

def candidatos_equilibrados(contagem_dict):
    """Retorna lista de candidatos dentro do limite de diferença de 1 plantão"""
    valores = list(contagem_dict.values())
    if not valores:
        return contagem_dict.items()
    minimo = min(valores)
    maximo_permitido = minimo + 1
    return [(c, n) for c, n in contagem_dict.items() if n <= maximo_permitido]

# -----------------------------
# Preenche lacunas diurnas de fim de semana
# -----------------------------
for row, data in lacunas_fds:
    candidatos_validos = candidatos_equilibrados(contagem_fs_diurno)
    candidatos_validos = sorted(candidatos_validos, key=lambda x: x[1])

    escolhido = None

    for codigo, _ in candidatos_validos:
        nome = mapa_codigos[codigo]
        if em_ferias(codigo, data):
            continue
        if trabalha_mesmo_dia(nome, data, escala):
            continue
        if not tem_folga(nome, data, escala):
            continue

        escolhido = (codigo, nome)
        break

    if escolhido:
        codigo, nome = escolhido
        ws.cell(row=row, column=col_diurno, value=nome)
        escala.append((row, data, nome, ""))  # adiciona ao histórico
        contagem_fs_diurno[codigo] += 1
    else:
        print(f"⚠️ Nenhum delegado disponível (com folga e equilíbrio) para {data.strftime('%d/%m/%Y')}")

# -----------------------------
# GERAR RESUMO DE PLANTÕES (MESMO MODELO DO ANTERIOR)
# -----------------------------
if "Resumo" in wb.sheetnames:
    del wb["Resumo"]
ws_resumo = wb.create_sheet("Resumo")

ws_resumo.append(["Delegado", "Diurno Semana", "Noturno Semana", "Diurno FimSemana", "Noturno FimSemana", "Total"])

# Inicializa contagem completa
contagem_total = {}

def contar(nome, categoria):
    if not nome:
        return
    if nome not in contagem_total:
        contagem_total[nome] = {
            "Diurno Semana": 0,
            "Noturno Semana": 0,
            "Diurno FimSemana": 0,
            "Noturno FimSemana": 0
        }
    contagem_total[nome][categoria] += 1

# Varre toda a escala e faz contagem atualizada
for _, data, diurno, noturno in escala:
    if not data:
        continue
    dia_semana = data.weekday()
    # Diurno
    if dia_semana < 5:
        contar(diurno, "Diurno Semana")
    else:
        contar(diurno, "Diurno FimSemana")
    # Noturno (sexta conta como fim de semana)
    if dia_semana <= 3:
        contar(noturno, "Noturno Semana")
    else:
        contar(noturno, "Noturno FimSemana")

# Monta resumo
for nome, dados in sorted(contagem_total.items()):
    total = sum(dados.values())
    ws_resumo.append([
        nome,
        dados["Diurno Semana"],
        dados["Noturno Semana"],
        dados["Diurno FimSemana"],
        dados["Noturno FimSemana"],
        total
    ])

# Ajusta largura das colunas
for col in ["A", "B", "C", "D", "E", "F"]:
    ws_resumo.column_dimensions[col].width = 25

# -----------------------------
# SALVAR RESULTADO
# -----------------------------
wb.save(saida_final)
print(f"✅ Escala final equilibrada salva em: {saida_final}")
print("📊 Resumo de plantões atualizado incluído na planilha.")
