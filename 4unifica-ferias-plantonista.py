import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# -----------------------------
# Função para transformar data
# -----------------------------
def parse_data(cell):
    if isinstance(cell, datetime):
        return cell.date()
    try:
        return datetime.strptime(str(cell), "%d/%m/%Y").date()
    except:
        return None

# -----------------------------
# Caminhos dos arquivos
# -----------------------------
arquivo_delegados = r"C:\Users\user\Desktop\Escala-de-Plant-o\delegados.xlsx"
arquivo_escala = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA 2º Semestre 2025.xlsx"
saida_final = r"C:\Users\user\Desktop\Escala-de-Plant-o\ESCALA_FINAL.xlsx"

# -----------------------------
# Carregar delegados e férias
# -----------------------------
df_delegados = pd.read_excel(arquivo_delegados)
df_delegados["Código"] = df_delegados["Código"].astype(str)  # força código como string

# Converte colunas de férias em datetime
for col in ["Inicio Férias 1", "Término Férias 1", "Inicio Férias 2", "Término Férias 2"]:
    df_delegados[col] = pd.to_datetime(df_delegados[col], errors="coerce")

# Dicionário de férias
ferias = {}  # {codigo: [(inicio1, fim1), (inicio2, fim2)]}
for _, row in df_delegados.iterrows():
    codigo = row["Código"]
    periodos = []
    if pd.notna(row["Inicio Férias 1"]) and pd.notna(row["Término Férias 1"]):
        periodos.append((row["Inicio Férias 1"].date(), row["Término Férias 1"].date()))
    if pd.notna(row["Inicio Férias 2"]) and pd.notna(row["Término Férias 2"]):
        periodos.append((row["Inicio Férias 2"].date(), row["Término Férias 2"].date()))
    if periodos:
        ferias[codigo] = periodos

# -----------------------------
# Carregar planilha da escala
# -----------------------------
wb = load_workbook(arquivo_escala)
ws = wb.active  # assume aba principal

col_data = 1   # coluna A
col_diurno = 3 # coluna C
col_noturno = 4 # coluna D

# -----------------------------
# Padrão de preenchimento
# -----------------------------
padrao = [("1", "2"), ("", "1"), ("2", ""), ("", ""), ("", "")]
indice = 0

# -----------------------------
# Preencher escala
# -----------------------------
for row in range(2, ws.max_row + 1):
    cell_data = ws.cell(row=row, column=col_data).value
    if not cell_data:
        continue

    data = parse_data(cell_data)
    if not data:
        continue

    # Pega códigos do padrão
    c_val, d_val = padrao[indice]
    indice = (indice + 1) % len(padrao)

    # Preenche provisoriamente
    ws.cell(row=row, column=col_diurno, value=c_val)
    ws.cell(row=row, column=col_noturno, value=d_val)

    # Verifica férias do diurno
    if c_val:
        for inicio, fim in ferias.get(c_val, []):
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_diurno, value="")  # apaga

    # Verifica férias do noturno
    if d_val:
        for inicio, fim in ferias.get(d_val, []):
            if (inicio - timedelta(days=1)) <= data <= (fim + timedelta(days=1)):
                ws.cell(row=row, column=col_noturno, value="")  # apaga

# -----------------------------
# Salvar planilha final
# -----------------------------
wb.save(saida_final)
print(f"✅ Escala final salva em: {saida_final}")
