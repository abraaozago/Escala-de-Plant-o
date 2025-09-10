import pandas as pd
from openpyxl import load_workbook

# Lê a planilha de delegados
delegados = pd.read_excel("delegados.xlsx", usecols="A:B", names=["Nome", "Codigo"])
mapa = dict(zip(delegados["Codigo"], delegados["Nome"]))

# Lê a planilha de escala usando openpyxl para manter a formatação
arquivo = "ESCALA 2º Semestre 2025.xlsx"
wb = load_workbook(arquivo)
ws = wb.active  # usa a primeira aba

# Substitui os códigos pelos nomes nas colunas C e D (índices 3 e 4)
for row in range(2, ws.max_row + 1):  # assume que a primeira linha é cabeçalho
    codigo_diurno = ws.cell(row=row, column=3).value
    codigo_noturno = ws.cell(row=row, column=4).value

    if codigo_diurno in mapa:
        ws.cell(row=row, column=3).value = mapa[codigo_diurno]
    if codigo_noturno in mapa:
        ws.cell(row=row, column=4).value = mapa[codigo_noturno]

# Salva em um novo arquivo mantendo formatação
wb.save("ESCALA_2º_Semestre_2025_com_nomes_formatada.xlsx")

print("✅ Planilha gerada com nomes mantendo formatação!")